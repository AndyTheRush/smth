from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os


folderName = 'test_dominigames'  #  folder name
folderPath = 'C:\\Users\\kisea\\Desktop\\'
Path = (folderPath + folderName + '\\')  #  files path
for name in list(os.listdir(path=Path)):
    if 'xlsx' in name:
        xlsxFileName = name.lower()
    elif 'plist' in name:
        plistFileName = name.lower()
    elif 'xml' in name:
        xmlFileName = name.lower()

wbExcel = load_workbook(Path + xlsxFileName)
ws = wbExcel.active

rowsNumber = 1
i = 1

while True:  # amount of rows with keys can be variable
    if ws['A' + str(i)].value is not None:
        rowsNumber += 1
        i += 1
    else:
        break

# lowercasing .xlsx file
for row in range(1, rowsNumber):
    for col in range(1, 5):
        char = get_column_letter(col)
        tempCellValue = ws[char + str(row)].value
        if type(tempCellValue) == int or tempCellValue == None:
            continue
        ws[char + str(row)] = tempCellValue.lower().encode('utf-8')

newInfoFile = []

# lowercasing .plist file
with open(Path + plistFileName, 'r') as infoFile:
    for row in infoFile:
        newRow = row.lower()
        newInfoFile.append(newRow)

with open(Path + plistFileName, 'w', encoding='utf-8') as infoFile:
    for row in newInfoFile:
        infoFile.write(row)

newDominiAP = []

# lowercasing .xml file
with open(Path + xmlFileName, 'r') as DominiAP:
    for row in DominiAP:
        newRow = row.lower()
        newDominiAP.append(newRow)

with open(Path + xmlFileName, 'w', encoding='utf-8') as DominiAP:
    for row in newDominiAP:
        DominiAP.write(row)

# saving file in lowercase
wbExcel.save(Path + xlsxFileName)
os.rename(Path + xlsxFileName, Path + xlsxFileName.lower())
os.rename(Path + plistFileName, Path + plistFileName.lower())
os.rename(Path + xmlFileName, Path + xmlFileName.lower())
os.rename(folderPath + folderName, folderPath + folderName.lower())
