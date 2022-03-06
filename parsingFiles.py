from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from itertools import islice
import os


class FileLowercasing:
    def __init__(self, folderPath, folderName):
        self.path = folderPath
        self.name = folderName
        self.fPath = (folderPath + folderName + '\\')

    def getting_files_names(self):  # asking out folder for file names
        get_xml_plist_files = []
        get_xlsxfiles = []
        for name in list(os.listdir(path=self.fPath)):  #  getting all file names
            if 'plist' in name or 'xml' in name:
                filename = name.lower()
                get_xml_plist_files.append(filename)
            elif 'xlsx' in name:
                filename = name.lower()
                get_xlsxfiles.append(filename)
        return get_xml_plist_files, get_xlsxfiles

    def files_lowercasing(self, get_xml_plist_files, get_xlsxfiles):  # func sets folder/all files/files content to lower case
        rowsnumber = 1
        valid = False
        for file in get_xml_plist_files:  #  lowercasing xml files
            newInfoFile = []
            with open(self.fPath + file, 'r') as infoFile:
                for row in infoFile:
                    newrow = row.lower()
                    newInfoFile.append(newrow)

            with open(self.fPath + file, 'w', encoding='utf-8') as infoFile:
                for row in newInfoFile:
                    infoFile.write(row)
            os.rename(self.fPath + file, self.fPath + file.lower())

        for file in get_xlsxfiles:
            wbExcel = load_workbook(self.fPath + file)
            ws = wbExcel.active
            while not valid:  # amount of rows with keys can be variable
                checkList = []
                [checkList.append(ws[chr(x + 65) + str(rowsnumber)].value) for x in range(4)]
                if any(checkList):
                    rowsnumber += 1
                else:
                    valid = True
            for row in range(1, rowsnumber):
                for col in range(1, 5):
                    char = get_column_letter(col)
                    tempCellValue = ws[char + str(row)].value
                    if type(tempCellValue) == int or tempCellValue == None:
                        continue
                    ws[char + str(row)] = tempCellValue.lower().encode('utf-8')
            os.rename(self.fPath + file, self.fPath + file.lower())

        os.rename(self.path + self.name, self.path + self.name.lower())
        return rowsnumber

    def get_wrong_keys(self, xlsx_file_name, get_xml_plist_files, rowsnumber):  # func to find wrong keys in xml/plist files if they exist in xlsx file
        needTuplesPlist = []
        needTuplesXml = []
        wbExcel = load_workbook(self.fPath + xlsx_file_name)
        ws = wbExcel.active
        for row in range(2, rowsnumber):
            for col in range(1, 5, 2):
                if col == (1 or 2):
                    char = get_column_letter(col)
                    tempCellValue = ws[char + str(row)].value
                    needTuplesPlist.append((tempCellValue, ws[get_column_letter(col + 1) + str(row)].value))
                else:
                    if ws[get_column_letter(col) + str(row)].value == None:
                        continue
                    char = get_column_letter(col)
                    tempCellValue = ws[char + str(row)].value
                    needTuplesXml.append((tempCellValue, ws[get_column_letter(col + 1) + str(row)].value))

        for file in get_xml_plist_files:
            rownum = 4
            if 'xml' in file:
                with open(self.fPath + file, encoding='utf-8') as currentfile:
                    for row in islice(currentfile, rownum, None):
                        if '<productid>' in row:
                            for tempTup in needTuplesXml:
                                tempKeyRow = row.strip().removeprefix('<productid>').removesuffix('</productid>')
                                if tempKeyRow in tempTup:
                                    currentfile.readline()
                                    myline = currentfile.readline().strip().removeprefix('<store_desc>').removesuffix('</store_desc>')
                                    if tempTup[1] in myline:
                                        continue
                                    else:
                                        print('WRONG VALUE, FOR KEY', tempTup[0], ' SHOULD BE', tempTup[1], 'IN', file)
                        rownum += 1
            else:
                rownum = 4
                with open(self.fPath + file, encoding='utf-8') as currentfile:
                    for row in islice(currentfile, rownum, None):
                        if '<key>' in row:
                            for tempTup in needTuplesPlist:
                                tempKeyRow = row.strip().removeprefix('<key>').removesuffix('</key>')
                                if tempKeyRow in tempTup:
                                    myLine = currentfile.readline().strip().removeprefix('<string>').removesuffix(
                                        '</string>')
                                    if myLine in tempTup:
                                        continue
                                    else:
                                        print('WRONG VALUE, FOR KEY', tempTup[0], ' SHOULD BE', tempTup[1], 'IN', file)
                        rownum += 1



myFile = FileLowercasing('C:\\Users\\kisea\\Desktop\\', 'test_dominigames')
myFile.files_lowercasing(myFile.getting_files_names()[0], myFile.getting_files_names()[1])
myFile.get_wrong_keys('DominiGames Test  Sheet.xlsx', myFile.getting_files_names()[0],
                      myFile.files_lowercasing(myFile.getting_files_names()[0], myFile.getting_files_names()[1]))  #  first arg is needed to concrete .xlsx file meanwhile second give us a portion of files with keys to compare
