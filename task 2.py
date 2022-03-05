from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from itertools import islice
import os


folderName = 'test_dominigames'  #  folder name
folderPath = 'C:\\Users\\kisea\\Desktop\\'
Path = (folderPath + folderName + '\\')  #  files path
get_files = []
for name in list(os.listdir(path=Path)):
    if 'xlsx' in name:
        xlsxFileName = name.lower()
        get_files.append(xlsxFileName)
    elif 'plist' in name:
        plistFileName = name.lower()
        get_files.append(plistFileName)
    elif 'xml' in name:
        xmlFileName = name.lower()
        get_files.append(xmlFileName)


wbExcel = load_workbook(Path + xlsxFileName)
ws = wbExcel.active

rowsNumber = 1
i = 1
needTuplesPlist = []
needTuplesXml = []

while True:  # amount of rows can be variable
    if ws['A' + str(i)].value is not None:
        rowsNumber += 1
        i += 1
    else:
        break

# making tuples to compare with .plist and .xml files
for row in range(2, rowsNumber):
    for col in range(1, 5, 2):
        if col == (1 or 2):
            char = get_column_letter(col)
            tempCellValue = ws[char + str(row)].value
            needTuplesPlist.append((tempCellValue, ws[get_column_letter(col + 1) + str(row)].value))
        else:
            if ws[get_column_letter(col) + str(row)].value == None:
                continue
            char = get_column_letter(col)
            tempCellValue = ws[char + str(row)].value
            needTuplesXml.append((tempCellValue, ws[get_column_letter(col + 1) + str(row)].value))

rowNumXml = 4
# finding wrong ones in .xml
with open(Path + xmlFileName, encoding='utf-8') as dominiiap:
    for row in islice(dominiiap, rowNumXml, None):
        if '<productid>' in row:
            for tempTup in needTuplesXml:
                tempKeyRow = row.strip().removeprefix('<productid>').removesuffix('</productid>')
                if tempKeyRow in tempTup:
                    dominiiap.readline()
                    myLine = dominiiap.readline().strip().removeprefix('<store_desc>').removesuffix('</store_desc>')
                    if tempTup[1] in myLine:
                        continue
                    else:
                        print('WRONG VALUE, FOR KEY', tempTup[0], ' SHOULD BE', tempTup[1], 'IN dominiiap.xml')
        rowNumXml += 1

# finding wrong ones in .plist
rowNumPlist = 4
with open(Path + plistFileName, encoding='utf-8') as infofile:
    for row in islice(infofile, rowNumPlist, None):
        if '<key>' in row:
            for tempTup in needTuplesPlist:
                tempKeyRow = row.strip().removeprefix('<key>').removesuffix('</key>')
                if tempKeyRow in tempTup:
                    myLine = infofile.readline().strip().removeprefix('<string>').removesuffix('</string>')
                    if myLine in tempTup:
                        continue
                    else:
                        print('WRONG VALUE, FOR KEY', tempTup[0], ' SHOULD BE', tempTup[1], 'IN infofile.plist')
        rowNumPlist += 1
